from openpyxl import Workbook, load_workbook
from Coordinate import Coordinate

def create_workbook(path: str, dataA: list[Coordinate], dataB: list[Coordinate]): 
    # create new workbook object
    workbook = Workbook()
    sheet = workbook.active

    # headers for procedure A
    sheet["A1"] = "Procedure A"
    sheet["A2"] = "Time t"
    sheet["B2"] = "Temperature Deg C"
    sheet["C2"] = "Estimated Slope"

    # headers for prodedure B
    sheet["E1"] = "Procedure B"
    sheet["E2"] = "Time t"
    sheet["F2"] = "Temperature Deg C"
    sheet["G2"] = "Estimated Slope"

    # populate cells for procedure A data
    for i, c in enumerate(dataA):
        if i != 0 and i != len(dataA) - 1:
            sheet[f"A{i + 2}"] = c.x
            sheet[f"B{i + 2}"] = c.y
            sheet[f"C{i + 2}"] = (c.get_slope(dataA[i + 1]) + c.get_slope(dataA[i - 1], True)) / 2

    # populate cells for procedure B data
    for i, c in enumerate(dataB):
        if i != 0 and i != len(dataB) - 1:

            sheet[f"E{i + 2}"] = c.x
            sheet[f"F{i + 2}"] = c.y
            sheet[f"G{i + 2}"] = (c.get_slope(dataB[i + 1]) + c.get_slope(dataB[i - 1], True)) / 2
    # saves workbook under given name
    workbook.save(path)

# only runs if targeted by python3 executable
if __name__ == "__main__":
    
    # initialize incoming data
    rawBook: Workbook = load_workbook("incData.xlsx")
    # choose Raw Data sheet
    worksheet = rawBook["Raw Data"]

    # grab data for procedure A
    pointsA: list[Coordinate] = []
    for t in range(6, 112):
        point = Coordinate(worksheet[f"C{t}"].value,
                           worksheet[f"D{t}"].value)
        pointsA.append(point)

    # grab data for procedure B
    pointsB: list[Coordinate] = []
    for t in range(6, 283):
        point = Coordinate(worksheet[f"G{t}"].value,
                           worksheet[f"H{t}"].value)
        pointsB.append(point)

    # creates workbook for slope data to go into
    create_workbook("dataWithSlopes.xlsx", pointsA, pointsB)
